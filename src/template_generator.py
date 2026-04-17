"""
Template Generator — creates the Excel IRO input template.

Sheets:
    1. Anleitung       — methodology instructions (German)
    2. IRO_Eingabe     — pre-filled IRO table; user fills scoring columns
    3. ESRS_Referenz   — ESRS topic reference table
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Colour constants
GREEN_HEADER = "C6EFCE"
GREEN_DARK = "375623"
BLUE_HEADER = "BDD7EE"
GREY_LOCKED = "F2F2F2"
ORANGE_HEADER = "FCE4D6"
WHITE = "FFFFFF"

CONFIG_PATH = os.path.join(os.path.dirname(__file__), "..", "config", "esrs_topics.json")


def _load_config() -> dict:
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def _thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_font(color=GREEN_DARK) -> Font:
    return Font(bold=True, color=color, size=10)


def _build_instructions(ws) -> None:
    ws.title = "Anleitung"
    ws.sheet_view.showGridLines = False

    # Title
    ws["B2"] = "Doppelte Wesentlichkeitsanalyse — Bedienungsanleitung"
    ws["B2"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("B2:I2")

    # Intro text
    intro = [
        ("B4", "Zweck dieses Tools"),
        ("B5", (
            "Dieses Excel-Template unterstützt Sie bei der Durchführung einer CSRD-konformen "
            "Doppelten Wesentlichkeitsanalyse gemäß EFRAG-Methodik. Es deckt 7 ESRS-Themen "
            "und 20 IROs (Impacts, Risks, Opportunities) ab."
        )),
        ("B7", "Schritt-für-Schritt-Anleitung"),
        ("B8", "1.  Wechseln Sie zum Tabellenblatt 'IRO_Eingabe'."),
        ("B9", "2.  Befüllen Sie die gelb markierten Spalten (E bis J) für jede IRO-Zeile."),
        ("B10", "3.  Speichern Sie die Datei und führen Sie anschließend den CLI-Befehl aus:"),
        ("B11", "    python main.py run --input <Pfad zur Datei> --company '<Firmenname>' --year <Jahr>"),
        ("B13", "Bewertungsskala (1–5)"),
    ]
    for cell_ref, text in intro:
        cell = ws[cell_ref]
        cell.value = text
        if cell_ref in ("B4", "B7", "B13"):
            cell.font = Font(bold=True, size=11, color="1F4E79")
        else:
            cell.font = Font(size=10)
        cell.alignment = Alignment(wrap_text=True)

    # Scale table
    scale_headers = ["Wert", "Auswirkungswesentlichkeit", "Finanzielle Wesentlichkeit"]
    scale_data = [
        [1, "Sehr gering / vernachlässigbar", "Sehr unwahrscheinlich / minimal"],
        [2, "Gering", "Unwahrscheinlich / gering"],
        [3, "Mittel (Schwellenwert)", "Möglich / mittel (Schwellenwert)"],
        [4, "Hoch", "Wahrscheinlich / hoch"],
        [5, "Sehr hoch / kritisch", "Sehr wahrscheinlich / kritisch"],
    ]
    header_fill = PatternFill("solid", fgColor=BLUE_HEADER)
    for col, header in enumerate(scale_headers, start=2):
        cell = ws.cell(row=14, column=col, value=header)
        cell.font = _header_font("1F4E79")
        cell.fill = header_fill
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row_data in enumerate(scale_data, start=15):
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True)
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="center")

    # Methodology note
    ws["B21"] = "Berechnungsformeln (EFRAG-Methodik)"
    ws["B21"].font = Font(bold=True, size=11, color="1F4E79")
    ws["B22"] = "Auswirkungsscore   = (Ausmaß × 0,4) + (Reichweite × 0,3) + (Behebbarkeit × 0,3)"
    ws["B23"] = "Finanzieller Score = (Eintrittswahrscheinlichkeit × Finanzielle Auswirkung) ÷ 5"
    ws["B24"] = "Wesentlich = Score ≥ 3,0 | Doppelt wesentlich = beide Scores ≥ 3,0"
    for row in range(22, 25):
        ws.cell(row=row, column=2).font = Font(size=10, name="Courier New")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 35
    for col in range(5, 12):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _build_iro_input(ws, topics: list) -> None:
    ws.title = "IRO_Eingabe"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E2"

    # Column headers
    headers = [
        "IRO-ID", "ESRS-Thema", "IRO-Typ", "Beschreibung",
        "Ausmaß\n(1–5)", "Reichweite\n(1–5)", "Behebbarkeit\n(1–5)",
        "Eintritts-\nwahrsch.\n(1–5)", "Fin. Auswirkung\n(1–5)", "Zeithorizont"
    ]
    header_fill = PatternFill("solid", fgColor=GREEN_HEADER)
    user_fill = PatternFill("solid", fgColor="FFF2CC")   # yellow for user-input cols
    locked_fill = PatternFill("solid", fgColor=GREY_LOCKED)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=9, color="375623")
        cell.fill = header_fill if col_idx <= 4 else user_fill
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 42

    # Data validation: scores 1–5
    dv_score = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="5",
        showErrorMessage=True,
        errorTitle="Ungültiger Wert",
        error="Bitte einen ganzzahligen Wert zwischen 1 und 5 eingeben.",
    )
    # Data validation: time horizon
    dv_horizon = DataValidation(
        type="list",
        formula1='"short,medium,long"',
        showErrorMessage=True,
        errorTitle="Ungültiger Wert",
        error="Bitte 'short', 'medium' oder 'long' auswählen.",
    )
    ws.add_data_validation(dv_score)
    ws.add_data_validation(dv_horizon)

    row = 2
    for topic in topics:
        topic_code = topic["code"]
        topic_name = topic["name"]
        for iro in topic["iros"]:
            cells = [
                (1, iro["iro_id"]),
                (2, topic_code),
                (3, iro["iro_type"]),
                (4, iro["description"]),
                (5, None), (6, None), (7, None),   # user input
                (8, None), (9, None),               # user input
                (10, None),                         # user input (horizon)
            ]
            for col_idx, value in cells:
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 4))
                if col_idx <= 4:
                    cell.fill = locked_fill
                    cell.font = Font(size=9)
                else:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    cell.font = Font(size=9)

            # Apply data validation to score columns (E–I) and horizon (J)
            score_range = f"E{row}:I{row}"
            horizon_cell = f"J{row}"
            dv_score.add(score_range)
            dv_horizon.add(horizon_cell)

            row += 1

    # Column widths
    col_widths = [10, 12, 12, 55, 10, 10, 10, 12, 14, 12]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 42


def _build_esrs_reference(ws) -> None:
    ws.title = "ESRS_Referenz"
    ws.sheet_view.showGridLines = False

    ws["B2"] = "ESRS-Themenreferenz"
    ws["B2"].font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells("B2:G2")

    headers = ["Code", "Thema", "Säule", "ESRS-Standard", "Beschreibung", "EFRAG-Referenz"]
    header_fill = PatternFill("solid", fgColor=BLUE_HEADER)
    for col_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, color="1F4E79")
        cell.fill = header_fill
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    reference_data = [
        ("E1", "Climate Change", "Environmental", "ESRS E1",
         "Treibhausgasemissionen, Klimaanpassung, Energiewende",
         "EFRAG ESRS E1 — Climate Change"),
        ("E2", "Pollution", "Environmental", "ESRS E2",
         "Luft-, Wasser- und Bodenverschmutzung, gefährliche Stoffe",
         "EFRAG ESRS E2 — Pollution"),
        ("E3", "Water & Marine Resources", "Environmental", "ESRS E3",
         "Wasserentnahme, Meeresressourcen, Gewässerqualität",
         "EFRAG ESRS E3 — Water and Marine Resources"),
        ("S1", "Own Workforce", "Social", "ESRS S1",
         "Arbeitsbedingungen, Gleichstellung, Sicherheit, Betriebsrat",
         "EFRAG ESRS S1 — Own Workforce"),
        ("S2", "Workers in Value Chain", "Social", "ESRS S2",
         "Lieferanten, Subunternehmer, Menschenrechte, LkSG",
         "EFRAG ESRS S2 — Workers in the Value Chain"),
        ("G1", "Business Conduct", "Governance", "ESRS G1",
         "Antikorruption, Compliance, Wettbewerb, Lobbying",
         "EFRAG ESRS G1 — Business Conduct"),
        ("G2", "Corporate Culture & Anti-Corruption", "Governance", "ESRS G2*",
         "Unternehmenskultur, Ethik, Whistleblowing, Transparenz",
         "EFRAG DMA Guidance — Governance Topics"),
    ]

    pillar_colors = {
        "Environmental": "E2EFDA",
        "Social": "DEEAF1",
        "Governance": "FFF2CC",
    }

    for row_idx, row_data in enumerate(reference_data, start=5):
        pillar_fill = PatternFill("solid", fgColor=pillar_colors.get(row_data[2], WHITE))
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _thin_border()
            cell.fill = pillar_fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[4].height = 28
    col_widths_ref = [4, 8, 28, 14, 14, 48, 36]
    for idx, width in enumerate(col_widths_ref, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws["B13"] = "* G2 ist kein eigenständiger ESRS-Standard; die Themen sind in der EFRAG-DMA-Guidance als eigene Kategorie geführt."
    ws["B13"].font = Font(size=8, italic=True, color="7F7F7F")
    ws.merge_cells("B13:G13")


def generate_template(output_path: str) -> None:
    """
    Generate the Excel IRO input template and save it to output_path.

    Args:
        output_path: File path for the generated .xlsx file.
    """
    config = _load_config()
    topics = config["topics"]

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ws_instructions = wb.create_sheet("Anleitung")
    ws_input = wb.create_sheet("IRO_Eingabe")
    ws_reference = wb.create_sheet("ESRS_Referenz")

    _build_instructions(ws_instructions)
    _build_iro_input(ws_input, topics)
    _build_esrs_reference(ws_reference)

    # Open on IRO_Eingabe by default
    wb.active = ws_input

    try:
        wb.save(output_path)
    except OSError as e:
        raise OSError(f"Could not save template to '{output_path}': {e}") from e

    print(f"[OK] Template gespeichert: {output_path}")
